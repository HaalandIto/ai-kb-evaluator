import os
for key in (
    "http_proxy", "https_proxy", "all_proxy",
    "HTTP_PROXY", "HTTPS_PROXY", "ALL_PROXY"
):
    os.environ.pop(key, None)

import requests                        
session = requests.Session()
session.trust_env = False               
import json
import re
import time
import logging
from logging.handlers import RotatingFileHandler
import pandas as pd
import argparse


from concurrent.futures import ThreadPoolExecutor, as_completed
parser = argparse.ArgumentParser(description="Dify QA & 本地合并工具")
parser.add_argument("--kb-name", default="共青团",
                    help="知识库名称前缀，默认 '知识库001'")
parser.add_argument("--max-workers", type=int, default=1,
                    help="最大并发线程数")
args = parser.parse_args()

KB_NAME = args.kb_name
MAX_WORKERS = args.max_workers


# ===== 基础配置 =====
DATASET_ID       = "4b5259b0-9ea4-4d2d-834a-b8bc4b486a4c"
DATASET_API_KEY  = "Bearer dataset-iP4pEJdWfH1dULoGarO28XLZ"
DATASET_API_BASE = "http://211.90.218.228:8888/v1"

CHAT_API_URL     = "http://211.90.218.228:8888/v1/chat-messages"
CHAT_API_KEY     = "Bearer app-3eIp7a3i1lIGRpDoByO1H22T"
AGENT_ID = "ZXH7u1xHgSmlgFzs"

DEESEEK_API_KEY  = "sk-fbd59078f53246cbaf7a2f7422c872e6"
DEESEEK_API_BASE = "https://api.deepseek.com/v1"

REPORT_DIR       = r"C:\Users\ASUS\Desktop\test654321"
os.makedirs(REPORT_DIR, exist_ok=True)

# ===== 日志配置 =====
LOG_PATH = os.path.join(REPORT_DIR, f"{KB_NAME}_dify_test.log")
logger = logging.getLogger("dify_test")
logger.setLevel(logging.DEBUG)
file_handler = RotatingFileHandler(LOG_PATH, maxBytes=5*1024*1024, backupCount=2, encoding="utf-8")
formatter = logging.Formatter("%(asctime)s %(levelname)s %(message)s")
file_handler.setFormatter(formatter)
logger.addHandler(file_handler)
console = logging.StreamHandler()
console.setFormatter(formatter)
logger.addHandler(console)

# ===== 函数 =====

def sanitize_filename(name: str) -> str:
    safe = "".join(c for c in name if c.isalnum() or c in " -_（）[]()")
    return f"{KB_NAME}_{(safe[:200] or 'report')}.txt"

def get_documents(dataset_id, retries=2):
    url = f"{DATASET_API_BASE}/datasets/{dataset_id}/documents"
    for attempt in range(retries):
        try:
            r = session.get(url,
                            headers={"Authorization": DATASET_API_KEY},
                            timeout=10,
                            proxies={})
            r.raise_for_status()
            return r.json()
        except Exception as e:
            logger.error(f"get_documents 失败: {e}")
            time.sleep(2)
    raise RuntimeError("无法获取文档列表")

def parse_file_info(inputs):
    if isinstance(inputs, str):
        inputs = json.loads(inputs)
    text = "\n".join(
        f"{item['id']}: {item['data_source_detail_dict']['upload_file']['name']}"
        for item in inputs.get("data", [])
    )
    matches = re.findall(r"([0-9a-f\-]{36}):\s*([^\n]+?)(?=(?:\s[0-9a-f\-]{36}:)|$)", text)
    return [m[0] for m in matches], [m[1] for m in matches]

def get_segments(dataset_id, file_id, retries=2):
    url = f"{DATASET_API_BASE}/datasets/{dataset_id}/documents/{file_id}/segments"
    params = {"page":1, "page_size":100}
    for attempt in range(retries):
        try:
            r = session.get(url, headers={"Authorization": DATASET_API_KEY}, params=params, timeout=20)
            r.raise_for_status()
            data = r.json().get("data", [])
            segs = data.get("items") if isinstance(data, dict) and "items" in data else data
            if segs:
                logger.info(f"文档 {file_id} 拿到 {len(segs)} 个 segment")
                return segs
        except Exception as e:
            logger.error(f"get_segments 失败: {e}")
        time.sleep(2)
    logger.warning(f"文档 {file_id} segments 为空")
    return []

POSSIBLE_TEXT_KEYS = ["text","content","segment_content","parsed_text","raw_text","body"]
def extract_segment_text(seg):
    for k in POSSIBLE_TEXT_KEYS:
        v = seg.get(k)
        if isinstance(v, str) and v.strip():
            return v.strip()
    return ""

QA_BLOCK_RE = re.compile(r"问题\s*(\d+)[：:]\s*(.*?)\n答案\s*\1[：:]\s*(.*?)(?:\n---|\Z)", re.DOTALL)
def parse_plain_qa(raw):
    qa = []
    for m in QA_BLOCK_RE.finditer(raw):
        q, a = m.group(2).strip(), m.group(3).strip()
        if q and a:
            qa.append({"question": q, "answer": a})
    return qa

# ===== LLM 生成与对比 =====
def generate_qa_pairs_via_llm(text, retries=2):
    if not text.strip():
        return []

    snippet = text[:8000]
    prompt = (
        "你是一个问答提炼专家，擅长从正式文档中提取精确要点，生成高质量、结构清晰的问题与标准答案。\n"
        "请根据下方正式文档内容，识别其关键信息，提炼出 10 个精准、代表性的问题及对应答案。\n\n"
        "【任务要求】\n"
        "1. 所有问题必须严格依据原始文本，不得编造、引申或泛化，并且尽量模仿老师和学生可能提出来的问题；\n"
        "2. 问题必须明确、完整，可独立理解，不得包含以下内容或特征：\n"
        "   - 模糊指代，如“文件编号”“文号”“文件名称”“本文件”“本部门”“本制度”“该条款”“上述规定”“以上内容”“相关事项”“这些文件”等；\n"
        "   - 需要依赖上下文才能理解的问题，或问题描述不自洽；\n"
        "   - 带有时间、地点、人物等不明确指向的模糊表达，如“近期”“有关人员”“相关部门”等；\n"
        "   - 带有多重歧义或可自由解释空间的问题；\n"
        "   - 使用“是否”“可以吗”等简单是非问法，但缺少完整情境的问题。\n"
        "   - 不得问文件或者办法的实施日期,修订某某制度的目的 以及印发日期这些主观性强或者无意义的问题。 \n"
        "3. 问题应聚焦于：制度条款、定量指标、积分标准、操作流程、角色职责、使用条件等；\n"
        "4. 答案必须：\n"
        "   - 紧扣题干提问要点，全面覆盖应答内容；\n"
        "   - 尽可能使用原文术语和措辞，避免同义改写；\n"
        "   - 涉及多项内容时，使用序号、破折号或列表分点表达；\n"
        "   - 所有数值、名词、限定条件必须与原文保持一致；\n"
        "5. 严禁生成以下内容：\n"
        "   - 文本总结或概述类内容；\n"
        "   - 推理性解释、评价性语言或跨段整合；\n"
        "   - 模糊、宽泛、不聚焦的问题；\n"
        "6. 输出格式必须严格为 JSON 数组，且满足以下结构（注意字段大小写）：\n"
        "[\n"
        '  { "Q": "问题1？", "A": "答案1。" },\n'
        '  { "Q": "问题2？", "A": "答案2。" },\n'
        '  { "Q": "问题3？", "A": "答案3。" }\n'
        "]\n"
        "7. 必须以 [ 开头、以 ] 结尾，禁止添加 markdown 代码块标记（如```json）或解释说明。\n"
        "8. 仅输出 JSON 内容，不添加任何额外提示、注释、换行符以外的文本。\n\n"
        "【请严格遵守以上格式，仅根据以下文档内容生成问答】：\n"
        f"{snippet}"
    )


    for _ in range(retries):
        try:
            r = session.post(
                f"{DEESEEK_API_BASE}/chat/completions",
                headers={
                    "Authorization": f"Bearer {DEESEEK_API_KEY}",
                    "Content-Type": "application/json"
                },
                json={
                    "model": "deepseek-chat",
                    "messages": [{"role": "user", "content": prompt}],
                    "temperature": 0
                },
                timeout=90
            )
            r.raise_for_status()
            raw = r.json()["choices"][0]["message"]["content"]

            m = re.search(r"\[\s*{.*?}\s*\]", raw, re.DOTALL)
            if m:
                arr = json.loads(m.group(0))
                qa = [
                    {
                        "question": o.get("Q") or o.get("question", ""),
                        "answer": o.get("A") or o.get("answer", "")
                    }
                    for o in arr
                    if o.get("Q") or o.get("question")
                ]
                if qa:
                    return qa[:10]

            qa_plain = parse_plain_qa(raw)
            if qa_plain:
                return qa_plain[:10]

        except Exception as e:
            logger.error(f"generate_qa_pairs_via_llm 失败: {e}")
            time.sleep(2)

    return []


def compare_qa(qa_list, answers, contexts, retries=3):
    def _compare_one(idx, qa, a_t, context_text):
        question_text = qa["question"]
        answer_x_text = qa["answer"]
        answer_y_text = a_t
        prompt = (
            f"你将收到多个问题的两个版本答案（如：答案 X-{idx} 和 答案 Y-{idx}），它们是对同一个问题的回答。\n"
           "你的任务是：\n"
            f"  逐对对比 X-{idx} 与 Y-{idx}；\n"
            f"  以 X-{idx} 为标准，判断 Y-{idx} 是否准确表达了核心内容；\n"
            f"  如果 Y-{idx} 有遗漏、理解错误或扩展不当，请在“理由”中指出；\n"
            "  对于语序、近义词或标点差异不扣分；\n"
            "  给出相似度（0–100）和评级（优秀 / 合格 / 不合格）。\n"
            "评分标准（满足正态分布，优秀：合格：不合格 ≈ 1：8：1）：\n"
            "  相似度 > 90：优秀（约 10%）\n"
            "  相似度 60–90：合格（约 80%）\n"
            "  相似度 < 60：不合格（约 10%）\n"
            "说明：分数尽量符合正态分布，避免集中在极端值。\n"
            "输出格式要求（必须严格遵守）：\n"
            "题号：\n"
            "题干：\n"
            f"答案 X-{idx}：\n"
            f"答案 Y-{idx}：\n"
            "理由：\n"
            "相似度：\n"
            "评级：\n"
            "每题之间用如下分隔线分隔---------（不可改动）\n"
            "其他说明：\n"
            "  不要输出解释、代码、引号、标记语言；\n"
            "  不要分析缺失题；\n"
            "  “理由”必须具体写明差异点；\n"
            "输出必须为纯文本格式，适合复制保存为 .txt 文件。\n\n"
            f"题干内容如下：\n{question_text}\n\n"
            f"参考原始文档内容如下：\n{context_text}\n\n"
            f"答案 X-{idx} 内容如下：\n{answer_x_text}\n\n"
            f"答案 Y-{idx} 内容如下：\n{answer_y_text}"
        )

        
        content = ""
        for attempt in range(1, retries + 1):
            try:
                r = session.post(
                    f"{DEESEEK_API_BASE}/chat/completions",
                    headers={"Authorization": f"Bearer {DEESEEK_API_KEY}", "Content-Type": "application/json"},
                    json={"model": "deepseek-chat", "messages":[{"role":"user","content":prompt}], "temperature":0},
                    timeout=180
                )
                r.raise_for_status()
                content = r.json()["choices"][0]["message"]["content"].strip()
                logger.info(f"compare_qa 第{idx}题第{attempt}次尝试成功")
                break
            except Exception as e:
                logger.error(f"compare_qa 第{idx}题第{attempt}次尝试失败: {e}")
                time.sleep(5)
        if not content:
            return (
                f"题号：{idx}\n"
                f"题干：{question_text}\n"
                f"答案 X-{idx}：{answer_x_text}\n"
                f"答案 Y-{idx}：{answer_y_text}\n"
                "理由：调用 Deepseek API 失败，未能完成对比。\n"
                "相似度：0\n"
                "评级：不合格\n"
                "---------"
            )
        else:
            return content

    results = [None] * len(qa_list)
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = [
            executor.submit(_compare_one, idx + 1, qa, ans, ctx)
            for idx, (qa, ans, ctx) in enumerate(zip(qa_list, answers, contexts))
        ]
        for fut in as_completed(futures):
            try:
                res = fut.result()
                m = re.search(r"(?:答案 X-|答案 Y-)(\d+)", res)
                if not m:
                    logger.warning(f"compare_qa 内容无法提取题号：{res[:100]}...")
                    continue
                idx = int(m.group(1)) - 1
                results[idx] = res
            except Exception as e:
                logger.error(f"compare_qa 结果解析失败: {e}")
    return results

def ask_agent_sse(questions, retries=3):
    """
    使用 SSE 模式并发向 Dify 智能体提问，返回每个问题的答案列表。
    - questions: 问题列表
    - retries: 每题最多重试次数
    """
    answers = [""] * len(questions)

    def _call(idx, q):
        logger.info(f"准备提问第 {idx + 1} 题：{q}")
        text = ""
        for attempt in range(1, retries + 1):
            try:
                r = session.post(
                    CHAT_API_URL,
                    headers={
                        "Authorization": CHAT_API_KEY,
                        "Content-Type": "application/json"
                    },
                    json={
                        "inputs": {},
                        "query": q,
                        "agent_id": AGENT_ID,
                        "response_mode": "streaming",
                        "conversation_id": "",
                        "user": "tester"
                    },

                    stream=True,
                    timeout=180  
                )
                r.raise_for_status()
                last = time.time()

                for line in r.iter_lines():
                    if not line:
                        if time.time() - last > 30:
                            logger.warning(
                                f"SSE 超时终止：第 {idx + 1} 题超过 30 秒无响应"
                            )
                            raise TimeoutError("SSE 无响应超时")
                        continue

                    last = time.time()
                    ch = line.decode("utf-8", errors="ignore")
                    if not ch.startswith("data:"):
                        continue

                    try:
                        data = json.loads(ch.replace("data:", "", 1).strip())
                        event_type = data.get("event", "")

                        if event_type in ("message", "message_replace"):
                            content_piece = data.get("content") or data.get("answer") or ""
                            if isinstance(content_piece, str) and content_piece.strip():
                                text += content_piece.strip()
                            else:
                                logger.debug(
                                    f"第 {idx + 1} 题空 message，继续等待后续流: {data}"
                                )

                        elif event_type == "message_end":
                            logger.debug(
                                f"第 {idx + 1} 题已接收到完整响应"
                            )

                        elif event_type == "error":
                            logger.error(
                                f"SSE 服务端返回 error 事件，触发重试: {data}"
                            )
                            raise RuntimeError("SSE 服务端 error，触发重试")

                        else:
                            logger.warning(
                                f"第 {idx + 1} 题接收到未知事件类型: {data}"
                            )

                    except Exception as e:
                        logger.error(
                            f"第 {idx + 1} 题解析 SSE 数据失败: {ch}, 错误: {e}"
                        )

                if text.strip():
                    logger.info(
                        f"第 {idx + 1} 题返回内容（前100字）：{text.strip()[:100]}"
                    )
                    break
                else:
                    logger.warning(
                        f"第 {idx + 1} 题尝试 {attempt}/{retries} 无有效返回内容"
                    )
                    time.sleep(2)

            except Exception as e:
                logger.error(
                    f"第 {idx + 1} 题第 {attempt}/{retries} 次请求失败: {e}"
                )
                time.sleep(2)

        else:
            logger.warning(
                f"第 {idx + 1} 题所有尝试均失败，最终返回空字符串"
            )

        return idx, text.strip()

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futures = [
            pool.submit(_call, i, q)
            for i, q in enumerate(questions)
        ]
        for fut in as_completed(futures):
            try:
                idx, ans = fut.result() 
                answers[idx] = ans
            except Exception as e:
                logger.error(f"ask_agent_sse 并发任务失败: {e}")

    for i in range(len(answers)):
        if not answers[i].strip():
            logger.warning(
                f"第 {i+1} 题答案缺失，使用默认填充值参与对比评分"
            )
            answers[i] = "【系统提示】：本题智能体未返回答案，无法回答。"

    return answers




def format_report(qa_list, answers, compares):
    return "\n".join([c for c in compares if isinstance(c, str)])

def main():
    docs = get_documents(DATASET_ID)
    ids, names = parse_file_info(docs)
    for fid, fname in zip(ids, names):
        logger.info(f"开始处理：{fname}")
        segs = get_segments(DATASET_ID, fid)
        texts = [extract_segment_text(s) for s in segs if extract_segment_text(s)]
        if not texts:
            logger.warning(f"{fname} 无有效文本，跳过")
            continue
        content = "\n\n".join(texts)
        qa = generate_qa_pairs_via_llm(content)
        if not qa:
            logger.warning(f"{fname} 无 QA，对比跳过")
            continue
        ans = ask_agent_sse([q["question"] for q in qa])
        contexts = [content] * len(ans)
        cmp_results = compare_qa(qa, ans, contexts)
        report = format_report(qa, ans, cmp_results)
        out_path = os.path.join(REPORT_DIR, sanitize_filename(fname))
        with open(out_path, "w", encoding="utf-8") as f:
            f.write(report)
        logger.info(f"已生成报告：{out_path}")

# ===== 本地合并 TXT 到 Excel =====
from openpyxl import load_workbook
from openpyxl.styles import Alignment

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

def combine_local(dir_path: str, output_name: str = None):
    filename = output_name or f"{''.join(c if c.isalnum() else '_' for c in KB_NAME)}_combined.xlsx"
    all_data = []
    txt_files = []

    for fname in sorted(os.listdir(dir_path)):
        if fname.endswith(".txt"):
            full_path = os.path.join(dir_path, fname)
            parsed = parse_txt_file(full_path)
            all_data.extend(parsed)
            txt_files.append(full_path)

    if not all_data:
        print("未找到符合格式的 txt 文件")
        return

    df = pd.DataFrame(all_data)
    df = df.drop_duplicates(subset=["文件名", "题号"], keep="first")

    try:
        df["题号"] = df["题号"].astype(int)
        df = df.sort_values(["文件名", "题号"])
    except:
        df = df.sort_values("文件名")

    out_path = os.path.join(dir_path, filename)
    df.to_excel(out_path, index=False)

    wb = load_workbook(out_path)
    ws = wb.active

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(name="等线", size=10)
        ws.row_dimensions[row[0].row].height = 30

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len * 1.2, 50)  

    wb.save(out_path)
    print(f"✅ 已保存合并表格到：{out_path}")

    deleted_count = 0
    for path in txt_files:
        try:
            os.remove(path)
            deleted_count += 1
        except Exception as e:
            print(f"❌ 删除失败：{path}，原因：{e}")

    print(f"🧹 已删除 {deleted_count} 个 .txt 文件")

from typing import List, Dict
import os
import re

def parse_txt_file(filepath: str) -> List[Dict[str, str]]:
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()

    blocks = content.split('---------')
    data = []

    for block in blocks:
        block = block.strip()
        if not block:
            continue
        m_qid = re.search(r"答案\s*[XY]-?(\d+)", block)
        qid = m_qid.group(1).strip() if m_qid else ""

        def extract_field(field_name):
            pattern = rf"(?s){field_name}\s*(?:X|Y)?-?\d*[:：]\s*([\s\S]*?)(?=\n(?:题干|答案 X|答案 Y|理由|相似度|评级)[:：]|$)"
            m = re.search(pattern, block, flags=re.S)
            return m.group(1).strip() if m else ""

        question    = extract_field("题干")
        answer_x    = extract_field("答案 X")
        answer_y    = extract_field("答案 Y")
        reason      = extract_field("理由")
        similarity  = extract_field("相似度")
        rating      = extract_field("评级")

        if not (answer_x or answer_y):
            continue  

        row = {
            "知识库": KB_NAME,
            "文件名": os.path.basename(filepath),
            "题号":    qid,
            "问答对":  question,
            "理由":    reason,
            "相似度":  similarity,
            "评级":    rating,
        }

        data.append(row)

    return data

def process_document(fid, fname):
    logger.info(f"开始处理：{fname}")
    segs = get_segments(DATASET_ID, fid)
    texts = [extract_segment_text(s) for s in segs if extract_segment_text(s)]
    if not texts:
        logger.warning(f"{fname} 无有效文本，跳过")
        return
    content = "\n\n".join(texts)
    qa = generate_qa_pairs_via_llm(content)
    if not qa:
        logger.warning(f"{fname} 无 QA，对比跳过")
        return
    ans = ask_agent_sse([q["question"] for q in qa])
    contexts = [content] * len(ans)  
    cmp_results = compare_qa(qa, ans, contexts) 
    report = format_report(qa, ans, cmp_results)
    out_path = os.path.join(REPORT_DIR, sanitize_filename(fname))
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(report)
    logger.info(f"已生成报告：{out_path}")


if __name__ == '__main__':
    docs = get_documents(DATASET_ID)
    ids, names = parse_file_info(docs)
    total_docs = len(ids)
    if total_docs == 0:
        logger.info("没有文档可处理，程序退出")
        exit(0)

    num_workers = min(MAX_WORKERS, total_docs)
    logger.info(f"共 {total_docs} 个文档，将启用 {num_workers} 个并发线程")

    with ThreadPoolExecutor(max_workers=num_workers) as executor:
        futures = [
            executor.submit(process_document, fid, fname)
            for fid, fname in zip(ids, names)
        ]
        for fut in as_completed(futures):
            try:
                fut.result()
            except Exception as e:
                logger.error(f"文档处理时出错：{e}")

    combine_local(REPORT_DIR)
